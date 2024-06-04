from django.shortcuts import get_object_or_404, redirect, render
# Lo necesario para "auth"
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test

# Paginación
from django.core.paginator import Paginator

# Nuestro formulario del CustomUser
from .forms import CustomUserCreationForm

# Respuestas http
from django.http import HttpResponse

from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy
import openpyxl # type: ignore


User = get_user_model()

# metodo que requiera que sea super user o admin
def admin_required(login_url=None):
    return user_passes_test(lambda usuario:usuario.is_superuser, login_url=login_url)

## @login_required # este metodo necesita que sea un usuario autenticado pero no necesariamente un Admin
def user_list(request):
    users = User.objects.all()
    paginator = Paginator(users, 5) # Mostrar 5 usuarios por página
    page_number = request.GET.get('page') # 2
    page_obj = paginator.get_page(page_number)
    return render(request, 'users/user_list.html', {'page_obj': page_obj })
    

# add_user require este decorador @admin_required(login_url='/login/')

def add_user(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('user_list')
        else:
            print('Form is not valid: ', form.errors)
    else:
        form = CustomUserCreationForm()
    return render(request, 'users/add_user.html', {'form': form })   
            
# export_users_to_excel require este decorador @admin_required(login_url='/login/')

def export_user_to_excel(request):
    # Obtiene todos los usuarios de la base de datos
    users = User.objects.all()

    # Establece la cabecera para indicar que la respuesta debe ser descargada como un archivo adjunto  users.xlsx
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers = {'Content-Disposition':'attachment; filename=users.xlsx'}
    )
    
    # Crea un nuevo libre de trabajo excel
    workbook = openpyxl.Workbook()
    
    # Obtenie la hoja de trabajo activa y le pone un titulo 'users'
    worksheet = workbook.active;
    worksheet.title = 'Users'
    
    columns = ['Username', 'Email', 'First Name', 'Last Name']
    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1):
        worksheet.cell(row=row_num, column=col_num, value=column_title)
        
    for user in users:
        row_num +=1;
        row = [user.username, user.email, user.first_name, user.last_name]
        for col_num, column_value in enumerate(row, 1):
            worksheet.cell(row=row_num, column=col_num, value=column_value)
    
    
    workbook.save(response)
    return response;
        

## Eliminar un usuario

def delete_user(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    user.delete()
    return redirect('user_list')
